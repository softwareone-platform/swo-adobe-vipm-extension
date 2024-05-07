import json
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from adobe_vipm.adobe.client import get_adobe_client
from adobe_vipm.adobe.config import get_config
from adobe_vipm.adobe.errors import AdobeAPIError, AuthorizationNotFoundError
from adobe_vipm.adobe.utils import join_phone_number
from adobe_vipm.adobe.validation import (
    is_valid_address_line_1_length,
    is_valid_address_line_2_length,
    is_valid_city_length,
    is_valid_company_name,
    is_valid_company_name_length,
    is_valid_country,
    is_valid_email,
    is_valid_first_last_name,
    is_valid_phone_number_length,
    is_valid_postal_code,
    is_valid_postal_code_length,
    is_valid_state_or_province,
)
from adobe_vipm.utils import find_first

COLUMNS = {
    "A": "authorization_uk",
    "B": "seller_uk",
    "C": "company_name",
    "D": "address_line_1",
    "E": "address_line_2",
    "F": "postal_code",
    "G": "city",
    "H": "region",
    "I": "country",
    "J": "phone_prefix",
    "K": "phone_number",
    "L": "contact_first_name",
    "M": "contact_last_name",
    "N": "contact_email",
    "O": "status",
    "P": "error_message",
}


class Command(BaseCommand):
    help = "Creates Resellers in Adobe VIP Marketplace"

    def success(self, message):
        self.stdout.write(self.style.SUCCESS(message), ending="\n")

    def info(self, message):
        self.stdout.write(message, ending="\n")

    def warning(self, message):
        self.stdout.write(self.style.WARNING(message), ending="\n")

    def error(self, message):
        self.stderr.write(self.style.ERROR(message), ending="\n")

    def add_arguments(self, parser):
        parser.add_argument(
            "infile",
            metavar="[RESELLERS EXCEL FILE]",
            nargs=1,
            help="Input Excel file with resellers data.",
        )

    def load_authorizations_data(self):
        return json.load(open(settings.EXTENSION_CONFIG["ADOBE_AUTHORIZATIONS_FILE"]))

    def validate_reseller_data(self, reseller_data):
        errors = self._validate_company_profile(reseller_data)
        errors.extend(self._validate_reseller_address(reseller_data))
        errors.extend(self._validate_reseller_contact(reseller_data))
        return errors

    def validate_input_file(self, workbook):
        if len(workbook.sheetnames) > 1:
            raise CommandError(
                f"Too many worksheet in the input file: {',' .join(workbook.sheetnames)}.",
            )
        ws = workbook.active

        for letter, expected_column in COLUMNS.items():
            actual_column = ws[f"{letter}1"].value
            if expected_column != actual_column:
                raise CommandError(
                    "Invalid input worksheet: expected column "
                    f"{letter} to be {expected_column}, found {actual_column}.",
                )

        if ws.max_row < 2:
            raise CommandError(
                "Invalid input worksheet: not enough data to process.",
            )

        return ws

    def load_row(self, row):
        return {
            COLUMNS[cel.column_letter]: str(cel.value) if cel.value else cel.value
            for cel in row
        }

    def prepare_reseller_data(self, row_data):
        return {
            "companyName": row_data["company_name"],
            "address": {
                "addressLine1": row_data["address_line_1"],
                "addressLine2": row_data["address_line_2"],
                "postalCode": row_data["postal_code"],
                "city": row_data["city"],
                "state": row_data["region"],
                "country": row_data["country"],
                "phone": row_data.get("phone_number", ""),
            },
            "contact": {
                "firstName": row_data["contact_first_name"],
                "lastName": row_data["contact_last_name"],
                "email": row_data["contact_email"],
                "phone": row_data.get("phone_number", ""),
            },
        }

    def add_reseller_to_authorization(
        self,
        authorizations_data,
        authorization_uk,
        seller_uk,
        reseller_id,
    ):
        auth_data = find_first(
            lambda auth: auth["authorization_uk"] == authorization_uk,
            authorizations_data["authorizations"],
        )
        auth_data["resellers"].append(
            {
                "id": reseller_id,
                "seller_uk": seller_uk,
            },
        )
        json.dump(
            authorizations_data,
            open(settings.EXTENSION_CONFIG["ADOBE_AUTHORIZATIONS_FILE"], "w"),
            indent=4,
        )

    def handle(self, *args, **options):
        adobe_config = get_config()
        adobe_client = get_adobe_client()
        authorizations_data = self.load_authorizations_data()

        excel_file = Path(options["infile"][0])

        if not excel_file.is_file():
            raise CommandError(f"Invalid Excel file provided: {excel_file}")

        workbook = load_workbook(excel_file)
        sheet = self.validate_input_file(workbook)
        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            row_data = self.load_row(row)
            if row_data["status"] == "OK":
                self.info(
                    f"Status is OK for {row_data['seller_uk']} - "
                    f"{row_data['company_name']} ({row_data['authorization_uk']}): skip it.",
                )
                continue
            authorization_uk = row_data["authorization_uk"]
            seller_uk = row_data["seller_uk"]
            authorization = None
            try:
                authorization = adobe_config.get_authorization(authorization_uk)
            except AuthorizationNotFoundError:
                sheet[f"O{idx}"].value = "KO"
                sheet[f"P{idx}"].value = "Authorization not found"
                workbook.save(excel_file)
                self.error(
                    f"Authorization not found for {row_data['seller_uk']} - "
                    f"{row_data['company_name']} ({row_data['authorization_uk']}).",
                )
                continue

            if adobe_config.reseller_exists(authorization, row_data["seller_uk"]):
                sheet[f"O{idx}"].value = "OK"
                sheet[f"P{idx}"].value = ""
                workbook.save(excel_file)
                self.warning(
                    f"Reseller {row_data['seller_uk']} - "
                    f"{row_data['company_name']} ({row_data['authorization_uk']}) "
                    "already exist.",
                )
                continue

            reseller_data = self.prepare_reseller_data(row_data)
            errors = self.validate_reseller_data(reseller_data)
            if errors:
                err_message = ", ".join(errors)
                sheet[f"O{idx}"].value = "KO"
                sheet[f"P{idx}"].value = err_message
                workbook.save(excel_file)
                self.error(
                    f"Error validating data of {row_data['seller_uk']} - "
                    f"{row_data['company_name']} ({row_data['authorization_uk']}): {err_message}.",
                )
                continue

            reseller_id = None
            try:
                reseller_id = adobe_client.create_reseller_account(
                    authorization_uk,
                    seller_uk,
                    reseller_data,
                )
            except AdobeAPIError as e:
                sheet[f"O{idx}"].value = "KO"
                sheet[f"P{idx}"].value = str(e)
                workbook.save(excel_file)
                self.error(
                    f"Error creating {row_data['seller_uk']} - "
                    f"{row_data['company_name']} ({row_data['authorization_uk']}): {str(e)}.",
                )
                continue

            self.add_reseller_to_authorization(
                authorizations_data,
                authorization_uk,
                seller_uk,
                reseller_id,
            )
            sheet[f"O{idx}"].value = "OK"
            workbook.save(excel_file)
            self.success(
                f"Reseller {row_data['seller_uk']} - "
                f"{row_data['company_name']} ({row_data['authorization_uk']}) "
                f"created: id is {reseller_id}.",
            )

    def _validate_company_profile(self, reseller_data):
        errors = []
        if not is_valid_company_name_length(reseller_data["companyName"]):
            errors.append("invalid company_name length.")

        if not is_valid_company_name(reseller_data["companyName"]):
            errors.append("invalid company_name")

        return errors

    def _validate_reseller_address(self, reseller_data):
        errors = []

        address = reseller_data["address"]
        country_code = address["country"]

        if not is_valid_country(country_code):
            errors.append("invalid country")
        else:
            if not is_valid_state_or_province(country_code, address["state"]):
                errors.append("invalid region")

            if not is_valid_postal_code(country_code, address["postalCode"]):
                errors.append("invalid postal_code")

        for field, validator_func, err_msg in (
            ("postalCode", is_valid_postal_code_length, "invalid postal:code length"),
            (
                "addressLine1",
                is_valid_address_line_1_length,
                "invalid address_line_1 length",
            ),
            (
                "addressLine2",
                is_valid_address_line_2_length,
                "invalid address_line_2 length",
            ),
            ("city", is_valid_city_length, "invalid city length"),
        ):
            if not validator_func(address[field]):
                errors.append(err_msg)

        return errors

    def _validate_reseller_contact(self, reseller_data):
        errors = []
        contact = reseller_data["contact"]

        if not is_valid_first_last_name(contact["firstName"]):
            errors.append("invalid contact_first_name")

        if not is_valid_first_last_name(contact["lastName"]):
            errors.append("invalid contact_last_name")

        if not is_valid_email(contact["email"]):
            errors.append("invalid contact_email")

        if contact.get("phone") and not is_valid_phone_number_length(
            join_phone_number(contact["phone"])
        ):
            errors.append("invalid phone_prefix/phone_number")

        return errors

import copy
import json
from io import StringIO

import pytest
from django.core.management import CommandError, call_command
from openpyxl import Workbook, load_workbook

from adobe_vipm.adobe.errors import AdobeAPIError
from adobe_vipm.management.commands.create_resellers import COLUMNS, Command

pytestmark = pytest.mark.usefixtures("mock_adobe_config")


def test_invalid_argument(mocker, adobe_authorizations_file, tmp_path):
    mocker.patch("adobe_vipm.management.commands.create_resellers.get_adobe_client")
    mocker.patch.object(
        Command,
        "load_authorizations_data",
        return_value=adobe_authorizations_file,
    )
    with pytest.raises(CommandError) as pr:
        call_command(
            "create_resellers",
            tmp_path.absolute(),
        )

    assert str(pr.value) == f"Invalid Excel file provided: {tmp_path.absolute()}"


def test_invalid_file_too_many_sheets(mocker, adobe_authorizations_file, tmp_path):
    mocker.patch("adobe_vipm.management.commands.create_resellers.get_adobe_client")
    wb = Workbook()
    wb.create_sheet("second")
    wb.save(tmp_path / "test.xlsx")

    mocker.patch.object(
        Command,
        "load_authorizations_data",
        return_value=adobe_authorizations_file,
    )

    with pytest.raises(CommandError) as pr:
        call_command(
            "create_resellers",
            tmp_path / "test.xlsx",
        )

    assert str(pr.value) == f"Too many worksheet in the input file: {','.join(wb.sheetnames)}."


def test_invalid_file_invalid_columns(mocker, adobe_authorizations_file, tmp_path):
    mocker.patch("adobe_vipm.management.commands.create_resellers.get_adobe_client")
    wb = Workbook()
    ws = wb.active
    ws["A1"].value = "Invalid column"
    wb.save(tmp_path / "test.xlsx")

    mocker.patch.object(
        Command,
        "load_authorizations_data",
        return_value=adobe_authorizations_file,
    )

    with pytest.raises(CommandError) as pr:
        call_command(
            "create_resellers",
            tmp_path / "test.xlsx",
        )

    assert str(pr.value) == (
        "Invalid input worksheet: expected column A to be authorization_uk, found Invalid column."
    )


def test_invalid_file_no_data(mocker, adobe_authorizations_file, tmp_path):
    mocker.patch("adobe_vipm.management.commands.create_resellers.get_adobe_client")
    wb = Workbook()
    ws = wb.active

    for letter, column in COLUMNS.items():
        ws[f"{letter}1"].value = column

    wb.save(tmp_path / "test.xlsx")

    mocker.patch.object(
        Command,
        "load_authorizations_data",
        return_value=adobe_authorizations_file,
    )

    with pytest.raises(CommandError) as pr:
        call_command(
            "create_resellers",
            tmp_path / "test.xlsx",
        )

    assert str(pr.value) == ("Invalid input worksheet: not enough data to process.")


def test_skip_processed(mocker, adobe_authorizations_file, tmp_path):
    mocker.patch("adobe_vipm.management.commands.create_resellers.get_adobe_client")
    wb = Workbook()
    ws = wb.active

    for letter, column in COLUMNS.items():
        ws[f"{letter}1"].value = column
        if letter not in {"O", "P"}:
            ws[f"{letter}2"].value = f"row_1_{column}"

    ws["O2"].value = "OK"
    wb.save(tmp_path / "test.xlsx")

    mocker.patch.object(
        Command,
        "load_authorizations_data",
        return_value=adobe_authorizations_file,
    )

    out = StringIO()

    call_command(
        "create_resellers",
        tmp_path / "test.xlsx",
        "--no-color",
        stdout=out,
    )

    assert "Status is OK for" in out.getvalue()


def test_authorization_not_found(mocker, adobe_authorizations_file, tmp_path):
    mocker.patch("adobe_vipm.management.commands.create_resellers.get_adobe_client")
    wb = Workbook()
    ws = wb.active

    for letter, column in COLUMNS.items():
        ws[f"{letter}1"].value = column
        if letter not in {"O", "P"}:
            ws[f"{letter}2"].value = f"row_1_{column}"

    wb.save(tmp_path / "test.xlsx")

    mocker.patch.object(
        Command,
        "load_authorizations_data",
        return_value=adobe_authorizations_file,
    )

    err = StringIO()

    call_command(
        "create_resellers",
        tmp_path / "test.xlsx",
        "--no-color",
        stderr=err,
    )

    wb = load_workbook(tmp_path / "test.xlsx")
    ws = wb.active

    assert "Authorization not found for " in err.getvalue()
    assert ws["O2"].value == "KO"
    assert ws["P2"].value == "Authorization not found"


def test_reseller_exists(mocker, settings, adobe_authorizations_file, tmp_path):
    settings.EXTENSION_CONFIG = {"ADOBE_AUTHORIZATIONS_FILE": "/path/to/authorizations.json"}
    mocker.patch("adobe_vipm.management.commands.create_resellers.get_adobe_client")
    mocker.patch(
        "adobe_vipm.management.commands.create_resellers.Path.open",
        mocker.mock_open(read_data=json.dumps(adobe_authorizations_file)),
    )
    authorization = adobe_authorizations_file["authorizations"][0]
    authorization_uk = authorization["authorization_uk"]
    seller_uk = authorization["resellers"][0]["seller_uk"]

    wb = Workbook()
    ws = wb.active

    for letter, column in COLUMNS.items():
        ws[f"{letter}1"].value = column
        if column == "authorization_uk":
            ws[f"{letter}2"].value = authorization_uk
            continue
        if column == "seller_uk":
            ws[f"{letter}2"].value = seller_uk
            continue
        if letter not in {"O", "P"}:
            ws[f"{letter}2"].value = f"row_1_{column}"

    wb.save(tmp_path / "test.xlsx")

    out = StringIO()

    call_command(
        "create_resellers",
        tmp_path / "test.xlsx",
        "--no-color",
        stdout=out,
    )

    wb = load_workbook(tmp_path / "test.xlsx")
    ws = wb.active

    assert "already exist." in out.getvalue()
    assert ws["O2"].value == "OK"
    assert ws["P2"].value is None


def test_reseller_create_ok(
    mocker, mock_adobe_client, settings, adobe_authorizations_file, tmp_path, reseller_data
):
    settings.EXTENSION_CONFIG = {"ADOBE_AUTHORIZATIONS_FILE": "/path/to/authorizations.json"}
    authorization = adobe_authorizations_file["authorizations"][0]
    authorization_uk = authorization["authorization_uk"]

    wb = Workbook()
    ws = wb.active

    for letter, column in COLUMNS.items():
        ws[f"{letter}1"].value = column
        if column == "authorization_uk":
            ws[f"{letter}2"].value = authorization_uk
            continue
        if column == "seller_uk":
            ws[f"{letter}2"].value = "another_seller_uk"
            continue
        if letter not in {"O", "P"}:
            ws[f"{letter}2"].value = f"row_1_{column}"

    wb.save(tmp_path / "test.xlsx")

    mocker.patch.object(
        Command,
        "load_authorizations_data",
        return_value=adobe_authorizations_file,
    )

    mocker.patch.object(
        Command,
        "prepare_reseller_data",
        return_value=reseller_data,
    )

    new_auth = copy.copy(adobe_authorizations_file)
    new_auth["authorizations"][0]["resellers"].append(
        {
            "seller_uk": "another_seller_uk",
            "id": "adobe-reseller-id",
        },
    )
    mocked_fobj = mocker.MagicMock()
    mocked_open = mocker.patch(
        "adobe_vipm.management.commands.create_resellers.Path.open",
        return_value=mocked_fobj,
    )
    mocked_dump = mocker.patch("adobe_vipm.management.commands.create_resellers.json.dump")

    mock_adobe_client.create_reseller_account.return_value = {"resellerId": "adobe-reseller-id"}
    out = StringIO()

    call_command(
        "create_resellers",
        tmp_path / "test.xlsx",
        "--no-color",
        stdout=out,
    )
    wb = load_workbook(tmp_path / "test.xlsx")
    ws = wb.active

    assert ws["O2"].value == "OK"
    assert ws["P2"].value is None

    mocked_dump.assert_called_once_with(
        new_auth,
        mocked_fobj.__enter__.return_value,
        indent=4,
    )
    mocked_open.assert_called_once_with("w", encoding="utf-8")


def test_api_error(
    mocker, mock_adobe_client, adobe_authorizations_file, tmp_path, adobe_api_error_factory
):
    authorization = adobe_authorizations_file["authorizations"][0]
    authorization_uk = authorization["authorization_uk"]

    wb = Workbook()
    ws = wb.active

    for letter, column in COLUMNS.items():
        ws[f"{letter}1"].value = column
        if column == "authorization_uk":
            ws[f"{letter}2"].value = authorization_uk
            continue
        if column == "seller_uk":
            ws[f"{letter}2"].value = "another_seller_uk"
            continue
        if letter not in {"O", "P"}:
            ws[f"{letter}2"].value = f"row_1_{column}"

    wb.save(tmp_path / "test.xlsx")
    mocker.patch.object(Command, "load_authorizations_data", return_value=adobe_authorizations_file)
    mocker.patch.object(Command, "validate_reseller_data", return_value=[])
    api_error = AdobeAPIError(
        400,
        adobe_api_error_factory(code="7777", message="Test Api Error"),
    )
    mock_adobe_client.create_reseller_account.side_effect = api_error
    out = StringIO()

    call_command(
        "create_resellers",
        tmp_path / "test.xlsx",
        "--no-color",
        stdout=out,
    )
    wb = load_workbook(tmp_path / "test.xlsx")
    ws = wb.active

    assert ws["O2"].value == "KO"
    assert ws["P2"].value == str(api_error)


def test_validation_errors_report(mocker, adobe_authorizations_file, tmp_path):
    mocker.patch("adobe_vipm.management.commands.create_resellers.get_adobe_client")
    authorization = adobe_authorizations_file["authorizations"][0]
    authorization_uk = authorization["authorization_uk"]

    wb = Workbook()
    ws = wb.active

    for letter, column in COLUMNS.items():
        ws[f"{letter}1"].value = column
        if column == "authorization_uk":
            ws[f"{letter}2"].value = authorization_uk
            continue
        if column == "seller_uk":
            ws[f"{letter}2"].value = "another_seller_uk"
            continue
        if letter not in {"O", "P"}:
            ws[f"{letter}2"].value = f"row_1_{column}"

    wb.save(tmp_path / "test.xlsx")

    mocker.patch.object(
        Command,
        "load_authorizations_data",
        return_value=adobe_authorizations_file,
    )

    mocker.patch.object(
        Command,
        "validate_reseller_data",
        return_value=[
            "error1",
            "error2",
        ],
    )

    out = StringIO()

    call_command(
        "create_resellers",
        tmp_path / "test.xlsx",
        "--no-color",
        stdout=out,
    )
    wb = load_workbook(tmp_path / "test.xlsx")
    ws = wb.active

    assert ws["O2"].value == "KO"
    assert ws["P2"].value == "error1, error2"


@pytest.mark.parametrize(
    ("field", "value"),
    [
        ("companyName", ""),
        ("companyName", "Euro € Company"),
        (
            "address",
            {
                "country": "ES",
                "state": "B",
                "city": "Barcelona",
                "addressLine1": "Plaza Catalunya 1",
                "addressLine2": "1o 1a",
                "postCode": "08001",
            },
        ),
        (
            "address",
            {
                "country": "US",
                "state": "ZZ",
                "city": "San Jose",
                "addressLine1": "3601 Lyon St",
                "addressLine2": "",
                "postCode": "94123",
            },
        ),
        (
            "address",
            {
                "country": "US",
                "state": "CA",
                "city": "San Jose",
                "addressLine1": "3601 Lyon St",
                "addressLine2": "",
                "postCode": "9412312",
            },
        ),
        (
            "address",
            {
                "country": "VU",
                "state": "TOB",
                "city": "Lalala",
                "addressLine1": "Blah blah",
                "addressLine2": "",
                "postCode": "9" * 41,
            },
        ),
        (
            "address",
            {
                "country": "VU",
                "state": "TOB",
                "city": "C" * 41,
                "addressLine1": "1" * 61,
                "addressLine2": "2" * 61,
                "postCode": "",
            },
        ),
        (
            "contact",
            {
                "firstName": "First N@m€",
                "lastName": "Last Name",
                "email": "test@example.com",
            },
        ),
        (
            "contact",
            {
                "firstName": "First Name",
                "lastName": "L@ast N@m€",
                "email": "test@example.com",
            },
        ),
        (
            "contact",
            {
                "firstName": "Test",
                "lastName": "Test",
                "email": "not-an-email",
            },
        ),
        (
            "contact",
            {
                "firstName": "First Name",
                "lastName": "Last Name",
                "email": "test@example.com",
                "phone": {
                    "prefix": "+1",
                    "number": "4082954078" * 5,
                },
            },
        ),
    ],
)
def test_validation_errors(
    mocker, adobe_authorizations_file, tmp_path, reseller_data, field, value
):
    mocker.patch("adobe_vipm.management.commands.create_resellers.get_adobe_client")
    authorization = adobe_authorizations_file["authorizations"][0]
    authorization_uk = authorization["authorization_uk"]

    modified_reseller_data = copy.copy(reseller_data)
    modified_reseller_data[field] = value

    wb = Workbook()
    ws = wb.active

    for letter, column in COLUMNS.items():
        ws[f"{letter}1"].value = column
        if column == "authorization_uk":
            ws[f"{letter}2"].value = authorization_uk
            continue
        if column == "seller_uk":
            ws[f"{letter}2"].value = "another_seller_uk"
            continue
        if letter not in {"O", "P"}:
            ws[f"{letter}2"].value = f"row_1_{column}"

    wb.save(tmp_path / "test.xlsx")

    mocker.patch.object(
        Command,
        "load_authorizations_data",
        return_value=adobe_authorizations_file,
    )

    mocker.patch.object(
        Command,
        "prepare_reseller_data",
        return_value=modified_reseller_data,
    )

    out = StringIO()

    call_command(
        "create_resellers",
        tmp_path / "test.xlsx",
        "--no-color",
        stdout=out,
    )
    wb = load_workbook(tmp_path / "test.xlsx")
    ws = wb.active

    assert ws["O2"].value == "KO"
    assert ws["P2"].value.startswith("invalid")
